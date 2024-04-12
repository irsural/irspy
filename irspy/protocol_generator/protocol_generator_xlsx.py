import shutil
from pathlib import Path
from typing import Iterable, Tuple, List
from zipfile import BadZipFile

from openpyxl.styles import Border, Side
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from irspy.protocol_generator.data_table import DataTable, DataCell
from irspy.protocol_generator.protocol_generator import ProtocolGenerator, ProtocolConfig
from irspy.protocol_generator.utils import handle_exceptions


class XlsxProtocolGenerator(ProtocolGenerator):
    def __init__(self) -> None:
        self.__last_detected_cell = None

    @handle_exceptions(BadZipFile, PermissionError, InvalidFileException)
    def generate(self, file: Path, config: ProtocolConfig) -> None:
            work_book = load_workbook(file)
            work_sheet = work_book.active

            while (ret := self.__find_cell(work_sheet, config.tag_map.keys())) is not None:
                tag, cell = ret
                value = config.tag_map[tag]
                if isinstance(value, DataTable):
                    self.__insert_table(work_sheet, cell, value)
                else:
                    cell.value = cell.value.replace(tag, value)

            work_book.save(file)

    def __resize_rows(self, work_sheet: Worksheet, start_row: int,
                      row_heights: List[int | None]) -> None:
        for row, row_height in enumerate(row_heights):
            if row_height is not None:
                work_sheet.row_dimensions[start_row + row].height = row_height

    def __resize_columns(self, work_sheet: Worksheet, start_column: int,
                         columns_widths: List[int | None]) -> None:
        for column, column_width in enumerate(columns_widths):
            if column_width is not None:
                letter = get_column_letter(start_column + column)
                work_sheet.column_dimensions[letter].width = column_width

    def __insert_table(self, work_sheet: Worksheet, cell: Cell, data_table: DataTable) -> None:
        cell_row = cell.row
        cell_column = cell.column
        work_sheet.insert_rows(cell.row, data_table.row_count - 1)

        for row_index, data_row in data_table.rows:
            for data_cell in data_row:
                curr_cell_row = data_cell.row + cell_row
                curr_cell_column = data_cell.column + cell_column

                start_cell = work_sheet.cell(row=curr_cell_row,
                                             column=curr_cell_column)
                start_cell.value = data_cell.value
                end_cell = work_sheet.cell(row=curr_cell_row + data_cell.row_span - 1,
                                           column=curr_cell_column + data_cell.column_span - 1)
                start_cell.alignment = Alignment(wrap_text=True, vertical='top')
                work_sheet.merge_cells(start_row=start_cell.row,
                                       start_column=start_cell.column,
                                       end_row=end_cell.row,
                                       end_column=end_cell.column)

        self.__add_border(work_sheet, cell_row, cell_column, data_table)
        self.__resize_rows(work_sheet, cell_row, data_table.xlsx_row_heights)
        self.__resize_columns(work_sheet, cell_column, data_table.xlsx_column_widths)

    def __add_border(self, work_sheet: Worksheet, row_start: int, column_start: int,
                     data_table: DataTable) -> None:
        if not data_table.add_border:
            return
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        for row in range(row_start, row_start + data_table.row_count):
            for column in range(column_start, column_start + data_table.column_count):
                work_sheet.cell(row=row, column=column).border = border

    def __find_cell(self, worksheet: Worksheet, texts: Iterable[str]) -> Tuple[str, Cell] | None:
        last_row, last_column = None, None
        if self.__last_detected_cell is not None:
            last_row, last_column = self.__last_detected_cell.row, self.__last_detected_cell.column
        for row in worksheet.iter_rows(min_row=last_row, min_col=last_column):
            for cell in row:
                for text in texts:
                    if cell.value is not None and text in str(cell.value):
                        self.__last_detected_cell = cell
                        return text, cell
        return None
